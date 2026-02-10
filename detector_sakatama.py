import openpyxl
from datetime import datetime

def extract_production_full_data(file_path, sheet_name):
    try:
        # 1. Load Workbook (data_only=True untuk mengambil nilai hasil rumus)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' tidak ditemukan.")
            return
        
        sheet = wb[sheet_name]

        # 2. Deteksi area "Production" menggunakan Merged Cells di Kolom B
        prod_min_row, prod_max_row = None, None
        for merged_range in sheet.merged_cells.ranges:
            top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            if top_left_cell.value == "Production":
                _, prod_min_row, _, prod_max_row = merged_range.bounds
                break

        if not prod_min_row:
            print("Area 'Production' tidak ditemukan.")
            return

        # 3. Definisikan Rentang Kolom (JK - XK) dan Pengecualian
        start_col = openpyxl.utils.column_index_from_string('JK')
        end_col = openpyxl.utils.column_index_from_string('XK')
        exclude_list = ["TOTAL CB", "TOTAL PCS", "TOTAL TON"]

        print(f"Mengekstrak data dari kolom JK sampai XK...")
        print(f"{'Tanggal':<12} | {'SKU':<12} | {'Deskripsi Produk':<45} | {'Qty':<8}")
        print("-" * 85)

        results = []

        # 4. Iterasi Kolom (Horizontal)
        for col_idx in range(start_col, end_col + 1):
            date_val = sheet.cell(row=17, column=col_idx).value
            
            # Formatting tanggal
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%d-%b-%Y')
            else:
                date_str = str(date_val) if date_val else "N/A"

            # 5. Iterasi Baris (Vertikal)
            for row_idx in range(prod_min_row, prod_max_row + 1):
                sku = sheet.cell(row=row_idx, column=1).value      # Kolom A
                product = sheet.cell(row=row_idx, column=3).value  # Kolom C
                qty = sheet.cell(row=row_idx, column=col_idx).value # Kolom JK-XK

                # --- FILTER: Cek Nama Produk & List Pengecualian ---
                if not product or any(x in str(product).upper() for x in exclude_list):
                    continue
                
                # --- FILTER: Hanya Qty > 0 ---
                if isinstance(qty, (int, float)) and qty > 0:
                    print(f"{date_str:<12} | {str(sku):<12} | {str(product)[:45]:<45} | {qty:<8}")
                    
                    results.append({
                        "tanggal": date_str,
                        "sku": sku,
                        "produk": product,
                        "qty": qty
                    })

        print("-" * 85)
        print(f"Berhasil menarik {len(results)} data produksi.")
        return results

    except Exception as e:
        print(f"Terjadi error: {e}")

# Jalankan
FILE_NAME = 'DPS_SAKATAMA.xlsx'
SHEET_TARGET = 'Demand vs Supply (DOH 15)'
final_data = extract_production_full_data(FILE_NAME, SHEET_TARGET)