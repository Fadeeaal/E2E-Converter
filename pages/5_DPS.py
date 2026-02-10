import io
import os
import re
from datetime import datetime

import openpyxl
import pandas as pd
import streamlit as st
from sqlalchemy import create_engine, text

st.set_page_config(page_title="DPS Cleaner Data", layout="wide")
st.title("DPS Cleaner Data")


# -------------------------
# Shared resources / helpers
# -------------------------
@st.cache_resource
def get_engine():
    p = st.secrets["postgres"]
    url = (
        f"postgresql+psycopg2://{p['user']}:{p['password']}"
        f"@{p['host']}:{p['port']}/{p['database']}"
        f"?sslmode=require"
    )
    return create_engine(url, pool_pre_ping=True)


engine = get_engine()


@st.cache_resource
def load_calendar_map():
    sql = text("SELECT cal_date, cal_week FROM calendar_cs")
    with engine.connect() as conn:
        df = pd.read_sql(sql, conn)
    df["cal_date"] = pd.to_datetime(df["cal_date"], errors="coerce").dt.date
    df = df.dropna(subset=["cal_date", "cal_week"])
    return dict(zip(df["cal_date"], df["cal_week"]))


@st.cache_resource
def load_master_data_map():
    """Mengambil referensi dari fg_master_data sebagai pengganti zcorin_converter"""
    sql = text(
        """
        SELECT sku_code, country, brand, sub_brand, category, big_category, house, pack_format, output, description
        FROM fg_master_data
    """
    )
    with engine.connect() as conn:
        df = pd.read_sql(sql, conn)

    df["sku_code"] = df["sku_code"].astype(str).str.strip()
    df = df.drop_duplicates(subset=["sku_code"])
    return df.set_index("sku_code").to_dict(orient="index")


CAL_MAP = load_calendar_map()
MASTER_MAP = load_master_data_map()


def norm(x) -> str:
    return str(x).strip().lower()


def sheet_has_line_header(excel_file, sheet_name: str, max_rows: int = 30) -> bool:
    preview = pd.read_excel(
        excel_file, sheet_name=sheet_name, header=None, nrows=max_rows, engine="openpyxl"
    )
    for r in range(len(preview)):
        vals = preview.iloc[r].tolist()
        vals = [v for v in vals if pd.notna(v)]
        if not vals:
            continue
        if "line" in [norm(v) for v in vals]:
            return True
    return False


def format_line_col_to_mon_yy(series: pd.Series) -> pd.Series:
    parsed = pd.to_datetime(series, errors="coerce")
    mask = parsed.notna()
    out = series.copy()
    out.loc[mask] = parsed.loc[mask].dt.strftime("%b-%y")
    return out


def calc_release_time(ts):
    if pd.isna(ts):
        return pd.NaT
    rt = ts + pd.Timedelta(days=5)
    if rt.weekday() == 5:
        rt += pd.Timedelta(days=2)
    elif rt.weekday() == 6:
        rt += pd.Timedelta(days=1)
    return rt


def detect_material_col(out: pd.DataFrame) -> str:
    cols = list(out.columns)
    col_map = {norm(c): c for c in cols}
    if "material" in col_map:
        return col_map["material"]
    if "sap" in col_map:
        return col_map["sap"]
    return cols[1] if len(cols) > 1 else cols[0]


def enrich_from_db(out: pd.DataFrame) -> pd.DataFrame:
    """Enrichment menggunakan mapping dari fg_master_data"""
    material_col = detect_material_col(out)
    keys = out[material_col].astype(str).str.strip()

    enrich_cols = [
        "country",
        "brand",
        "sub_brand",
        "category",
        "big_category",
        "house",
        "pack_format",
        "output",
    ]
    for c in enrich_cols:
        out[c] = keys.map(lambda k: MASTER_MAP.get(k, {}).get(c))
    return out


def round_FGH(out: pd.DataFrame) -> pd.DataFrame:
    """
    Round columns F,G,H from original excel selection A:H.
    In our selected out (A:H + O:P), FGH correspond to indices 5,6,7.
    """
    for idx in [5, 6, 7]:
        if idx < out.shape[1]:
            col = out.columns[idx]
            out[col] = pd.to_numeric(out[col], errors="coerce").round(0)
    return out


def filter_by_m0_m2(out: pd.DataFrame, month_set: set) -> pd.DataFrame:
    """
    Keep rows where Time_Finish month is in M0-M2 (based on parsed datetime).
    Assumption: out includes O:P (Time Start, Time_Finish) as last two cols.
    """
    time_start_col = out.columns[-2]
    time_finish_col = out.columns[-1]

    out[time_start_col] = pd.to_datetime(out[time_start_col], errors="coerce")
    out[time_finish_col] = pd.to_datetime(out[time_finish_col], errors="coerce")
    out = out[out[time_finish_col].dt.month.isin(month_set)].copy()
    out = out.sort_values(by=time_start_col, ascending=True)

    return out


def process_sheet(excel_file, sheet_name: str, month_set: set):
    if not sheet_has_line_header(excel_file, sheet_name):
        return None, "SKIP (no 'Line' header found)"

    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=0, engine="openpyxl")
    if df.shape[1] < 16:
        return None, "SKIP (not enough columns for A:H + O:P)"
    cols_idx = list(range(0, 8)) + list(range(14, 16))
    out = df.iloc[:, cols_idx].copy()
    out = out.dropna(how="all")

    line_col = out.columns[0]
    out[line_col] = format_line_col_to_mon_yy(out[line_col])
    out = round_FGH(out)
    out = filter_by_m0_m2(out, month_set)

    if out.empty:
        return None, "SKIP (no rows in selected months M0-M2)"
    time_finish_col = out.columns[-1]
    release_ts = out[time_finish_col].apply(calc_release_time)
    out["Release time"] = pd.to_datetime(release_ts, errors="coerce").dt.date
    out["Release wk"] = out["Release time"].map(CAL_MAP)

    out = enrich_from_db(out)
    # Drop 'machine_1' column if present
    if "machine_1" in out.columns:
        out = out.drop(columns=["machine_1"])

    return out, "OK"


def load_east_master_reference(engine) -> pd.DataFrame:
    """Mengambil semua data speed, size, dan kg_cb dari fg_master_data"""
    sql = text(
        """
        SELECT sku_code, line, kg_cb, size, speed, description, 
               country, brand, sub_brand, category, big_category, house, pack_format, output
        FROM fg_master_data
    """
    )
    with engine.connect() as conn:
        df = pd.read_sql(sql, conn)

    df["sku_code"] = df["sku_code"].astype(str).str.strip()
    df["line"] = df["line"].astype(str).str.strip().str.upper()
    return df


VALID_LINES = {"AB", "CD", "GH", "JK", "TU", "VW", "XY"}
DATE_ROW_IDX = 8
DATE_START_COL = 24
DATE_END_COL = 93
COL_MATERIAL = 5
COL_DESC = 6
COL_KG_CB = 9
COL_LINE = 10


def load_fg_master_data(engine) -> pd.DataFrame:
    """Mengambil semua referensi (enrichment & speed) dari fg_master_data"""
    sql = text(
        """
        SELECT 
            sku_code, line, description, pcs_cb, kg_cb, size, 
            country, brand, sub_brand, category, big_category, 
            house, region, speed, pack_format, output
        FROM fg_master_data
    """
    )
    with engine.connect() as conn:
        df = pd.read_sql(sql, conn)

    # Standarisasi kolom kunci
    df["sku_code"] = df["sku_code"].astype(str).str.strip()
    df["line"] = df["line"].astype(str).str.strip().str.upper()

    # Pastikan numerik
    df["speed"] = pd.to_numeric(df["speed"], errors="coerce")
    df["size"] = pd.to_numeric(df["size"], errors="coerce")
    df["kg_cb"] = pd.to_numeric(df["kg_cb"], errors="coerce")

    return df


def validate_east_sheet_format(raw: pd.DataFrame) -> tuple[bool, str]:
    """
    Validate if the sheet has the expected EAST format.
    Returns (is_valid, error_message)
    """
    if len(raw) <= DATE_ROW_IDX:
        return (
            False,
            f"Sheet has insufficient rows. Expected at least {DATE_ROW_IDX + 1} rows.",
        )

    if raw.shape[1] <= DATE_END_COL:
        return (
            False,
            f"Sheet has insufficient columns. Expected at least {DATE_END_COL + 1} columns.",
        )

    date_cells = raw.iloc[DATE_ROW_IDX, DATE_START_COL : DATE_END_COL + 1]
    dates = pd.to_datetime(date_cells, errors="coerce")
    valid_dates = dates.notna().sum()

    if valid_dates == 0:
        return False, "No valid date headers found in row 9 (columns Y to CP)."

    df_items = raw.copy()
    df_items[COL_LINE] = df_items[COL_LINE].astype(str).str.strip().str.upper()
    valid_line_rows = df_items[df_items[COL_LINE].isin(VALID_LINES)]

    if len(valid_line_rows) == 0:
        return (
            False,
            f"No rows found with valid Line values ({', '.join(sorted(VALID_LINES))}).",
        )

    return True, ""


def process_east_file(raw: pd.DataFrame, engine, month_set: set, cal_map: dict) -> dict:
    # 1) Load single source of truth
    master_ref = load_fg_master_data(engine).copy()

    # keep only needed cols + de-dup for stable merge
    needed_cols = [
        "sku_code",
        "line",
        "description",
        "kg_cb",
        "size",
        "speed",
        "country",
        "brand",
        "sub_brand",
        "category",
        "big_category",
        "house",
        "pack_format",
        "output",
    ]
    master_ref = master_ref[[c for c in needed_cols if c in master_ref.columns]].copy()
    master_ref["sku_code"] = master_ref["sku_code"].astype(str).str.strip()
    master_ref["line"] = master_ref["line"].astype(str).str.strip().str.upper()
    master_ref = master_ref.drop_duplicates(subset=["sku_code", "line"], keep="first")

    # 2) Detect valid date headers (row 9, cols Y..CP)
    date_cells = raw.iloc[DATE_ROW_IDX, DATE_START_COL : DATE_END_COL + 1]
    dates = pd.to_datetime(date_cells, errors="coerce")
    valid_date_mask = dates.notna()
    date_cols_idx = [
        DATE_START_COL + i for i, ok in enumerate(valid_date_mask.tolist()) if ok
    ]
    date_vals = dates[valid_date_mask].dt.date.tolist()

    # 3) Filter valid line rows
    df_items = raw.copy()
    df_items[COL_LINE] = df_items[COL_LINE].astype(str).str.strip().str.upper()
    df_items = df_items[df_items[COL_LINE].isin(VALID_LINES)].copy()

    # 4) Build wide then melt long
    keep_cols = [COL_MATERIAL, COL_DESC, COL_KG_CB, COL_LINE] + date_cols_idx
    df_wide = df_items.iloc[:, keep_cols].copy()
    df_wide.columns = ["Material", "Description", "Kg_TU", "Line"] + [
        str(d) for d in date_vals
    ]

    df_wide["Material"] = df_wide["Material"].astype(str).str.strip()
    df_wide["Description"] = df_wide["Description"].astype(str).str.strip()
    df_wide["Kg_TU"] = pd.to_numeric(df_wide["Kg_TU"], errors="coerce")
    df_wide["Line"] = df_wide["Line"].astype(str).str.strip().str.upper()

    # remove blank material rows
    df_wide = df_wide[
        (df_wide["Material"] != "") & (df_wide["Material"].str.lower() != "nan")
    ].copy()

    out = df_wide.melt(
        id_vars=["Material", "Description", "Kg_TU", "Line"],
        var_name="Date",
        value_name="Qty",
    )

    out["Date"] = pd.to_datetime(out["Date"], errors="coerce").dt.date
    out["Qty"] = pd.to_numeric(out["Qty"], errors="coerce")
    out = out[out["Qty"].fillna(0) > 0].copy()

    # 5) Merge enrichment + pack size + speed ONLY from fg_master_data
    out["Material"] = out["Material"].astype(str).str.strip()
    out["Line"] = out["Line"].astype(str).str.strip().str.upper()

    out = (
        out.merge(
            master_ref,
            how="left",
            left_on=["Material", "Line"],
            right_on=["sku_code", "line"],
        )
        .drop(columns=["sku_code", "line"], errors="ignore")
        .copy()
    )

    # Prefer sheet description, but fill missing from master
    if "description" in out.columns:
        out["Description"] = out["Description"].where(
            out["Description"].notna()
            & (out["Description"].astype(str).str.strip() != ""),
            out["description"],
        )
        out = out.drop(columns=["description"], errors="ignore")

    # Rename master fields to output names
    if "size" in out.columns:
        out = out.rename(columns={"size": "Pack Size"})
    else:
        out["Pack Size"] = None

    if "speed" in out.columns:
        out = out.rename(columns={"speed": "Speed"})
    else:
        out["Speed"] = None

    # Ensure numeric for computations
    out["Kg_TU"] = pd.to_numeric(out["Kg_TU"], errors="coerce")
    out["Speed"] = pd.to_numeric(out["Speed"], errors="coerce")

    # 6) Calculations
    out["Qty Bulk in KG"] = out["Qty"] * out["Kg_TU"]
    out["BIN"] = out["Qty Bulk in KG"] / 750

    # avoid division by zero
    out["Prod Hour"] = out.apply(
        lambda r: (r["Qty"] / r["Speed"])
        if pd.notna(r["Speed"]) and r["Speed"] not in [0, 0.0]
        else None,
        axis=1,
    )
    out["Days"] = out["Prod Hour"].apply(lambda x: (x / 24) if pd.notna(x) else None)

    # Remove duplicates at day/material/line/kg_tu level
    key_cols = ["Date", "Material", "Line", "Kg_TU"]
    out = out.drop_duplicates(subset=key_cols, keep="first")

    # 7) Build per-line schedules like your original code
    out = out.sort_values(["Date", "Line", "Material"], ascending=True)

    unique_lines = sorted(out["Line"].dropna().unique())
    line_dfs = {}

    for line in unique_lines:
        line_df = out[out["Line"] == line].copy()
        line_df["_orig_date"] = pd.to_datetime(line_df["Date"], errors="coerce")

        sorted_rows = []
        unique_dates = sorted(line_df["_orig_date"].unique())
        last_material_prev_day = None

        for d in unique_dates:
            day_data = line_df[line_df["_orig_date"] == d].copy()

            if last_material_prev_day is not None:
                priority_df = day_data[day_data["Material"] == last_material_prev_day]
                others_df = day_data[day_data["Material"] != last_material_prev_day].sort_values(
                    "Material", ascending=True
                )
                day_sorted = pd.concat([priority_df, others_df])
            else:
                day_sorted = day_data.sort_values("Material", ascending=True)

            if not day_sorted.empty:
                last_material_prev_day = day_sorted.iloc[-1]["Material"]

            sorted_rows.append(day_sorted)

        line_df = pd.concat(sorted_rows).reset_index(drop=True)

        # Date to Mon-YY string for final output
        if "Date" in line_df.columns:
            line_df["Date"] = pd.to_datetime(line_df["Date"], errors="coerce").dt.strftime(
                "%b-%y"
            )

        time_starts = []
        time_finishes = []

        for idx, row in line_df.iterrows():
            current_date_6am = pd.Timestamp(row["_orig_date"]) + pd.Timedelta(hours=6)

            if idx == 0:
                time_start = pd.Timestamp(row["_orig_date"]) + pd.Timedelta(hours=7)
            else:
                prev_finish = time_finishes[-1]
                time_start = prev_finish if prev_finish > current_date_6am else current_date_6am

            days_value = row["Days"] if pd.notna(row["Days"]) else 0
            time_finish = time_start + pd.Timedelta(days=days_value)

            time_starts.append(time_start)
            time_finishes.append(time_finish)

        line_df["Time Start"] = time_starts
        line_df["Time Finish"] = time_finishes

        # Filter by month set (M0..M2)
        line_df = line_df[line_df["Time Finish"].dt.month.isin(month_set)].copy()
        if line_df.empty:
            continue

        line_df = line_df.sort_values("Time Start", ascending=True).reset_index(drop=True)

        line_df["Release Time"] = line_df["Time Finish"].apply(calc_release_time)
        line_df["Release Time"] = pd.to_datetime(line_df["Release Time"], errors="coerce").dt.date
        line_df["Release wk"] = line_df["Release Time"].map(cal_map)

        # Final selection + rename like before
        final_cols_with_time = [
            "Date",
            "Material",
            "Description",
            "Pack Size",
            "Kg_TU",
            "Qty",
            "Qty Bulk in KG",
            "BIN",
            "Time Start",
            "Time Finish",
            "Release Time",
            "Release wk",
            "country",
            "brand",
            "sub_brand",
            "category",
            "big_category",
            "house",
            "pack_format",
        ]
        line_df = line_df[[c for c in final_cols_with_time if c in line_df.columns]].copy()

        line_df = line_df.rename(
            columns={
                "Material": "SAP Article",
                "Qty": "Qty (Ctn)",
                "Qty Bulk in KG": "Qty Bulk (kg)",
                "Release wk": "Release Week",
            }
        )

        # Drop helper
        line_df = line_df.drop(columns=["_orig_date"], errors="ignore")

        # Release Ident
        if "Release Time" in line_df.columns and "Release Week" in line_df.columns:

            def rel_ident_fmt(x):
                if pd.notna(x):
                    return f"{x.day}{x.month}{x.year}"
                return ""

            rel_ident = line_df["Release Time"].apply(rel_ident_fmt)
            idx = line_df.columns.get_loc("Release Week")
            line_df.insert(idx + 1, "Release Ident", rel_ident)

        # Ensure 'Line' exists
        if "Line" not in line_df.columns:
            line_df.insert(0, "Line", line)

        cols = ["Line"] + [c for c in TARGET_OUTPUT_COLS if c in line_df.columns]
        for extra in [
            "country",
            "brand",
            "sub_brand",
            "category",
            "big_category",
            "house",
            "pack_format",
        ]:
            if extra in line_df.columns and extra not in cols:
                cols.append(extra)

        line_dfs[line] = line_df[cols].copy()

    return line_dfs


def create_east_excel_download(line_dfs: dict) -> bytes:
    """Create Excel file with separate sheets per line and return as bytes."""
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for line in sorted(line_dfs.keys()):
            line_df = line_dfs[line].copy()
            # Ensure a Line column exists with the line name
            if "Line" not in line_df.columns:
                line_df.insert(0, "Line", line)
            line_df.to_excel(writer, sheet_name=f"Line_{line}", index=False)

        if line_dfs:
            # Concatenate and ensure 'Line' exists
            all_east_df = pd.concat(line_dfs.values(), ignore_index=True)
            if "Line" not in all_east_df.columns:
                # Try to infer line from sheet-level keys by adding a placeholder
                all_east_df.insert(0, "Line", None)
            all_east_df.to_excel(writer, sheet_name="All_East", index=False)

    output.seek(0)
    return output.getvalue()


# Desired final column order for both West and East outputs
TARGET_OUTPUT_COLS = [
    "Date",
    "SAP Article",
    "Description",
    "Pack Size",
    "Kg_TU",
    "Qty (Ctn)",
    "Qty Bulk (kg)",
    "BIN",
    "Time Start",
    "Time Finish",
    "Release Time",
    "Release Week",
    "Release Ident",
]


def ensure_output_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Return DataFrame with columns in TARGET_OUTPUT_COLS order.
    Case-insensitive mapping is used to find existing similar column names
    (e.g., 'Rel Ident' -> 'Release Ident', 'Qty (ctn)' -> 'Qty (Ctn)').
    Missing columns are created with None values.
    """
    existing_map = {c.lower().strip(): c for c in df.columns}
    out = pd.DataFrame()
    for t in TARGET_OUTPUT_COLS:
        key = t.lower().strip()
        found = existing_map.get(key)
        # common alias: 'rel ident' -> 'release ident'
        if not found and key == "release ident" and "rel ident" in existing_map:
            found = existing_map["rel ident"]
        if found:
            out[t] = df[found]
        else:
            out[t] = None
    return out


# -------------------------
# Sakatama processing helpers
# -------------------------
SAKATAMA_LINE = "A,B"
SAKATAMA_DATE_ROW = 17
SAKATAMA_START_COL = "JK"
SAKATAMA_END_COL = "XK"
SAKATAMA_EXCLUDE_LIST = ["TOTAL CB", "TOTAL PCS", "TOTAL TON"]


def extract_sakatama_production_data(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' tidak ditemukan.")
    sheet = wb[sheet_name]

    # Deteksi area "Production" menggunakan merged cells
    prod_min_row, prod_max_row = None, None
    for merged_range in sheet.merged_cells.ranges:
        top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
        if top_left_cell.value == "Production":
            prod_min_row = merged_range.min_row
            prod_max_row = merged_range.max_row
            break

    if not prod_min_row:
        raise ValueError("Area 'Production' tidak ditemukan.")

    start_col = openpyxl.utils.column_index_from_string(SAKATAMA_START_COL)
    end_col = openpyxl.utils.column_index_from_string(SAKATAMA_END_COL)

    rows = []

    for col_idx in range(start_col, end_col + 1):
        date_val = sheet.cell(row=SAKATAMA_DATE_ROW, column=col_idx).value
        parsed_date = pd.to_datetime(date_val, errors="coerce")
        if pd.isna(parsed_date):
            continue
        date_only = parsed_date.date()

        for row_idx in range(prod_min_row, prod_max_row + 1):
            sku = sheet.cell(row=row_idx, column=1).value
            product = sheet.cell(row=row_idx, column=3).value
            qty = sheet.cell(row=row_idx, column=col_idx).value

            if not product or any(x in str(product).upper() for x in SAKATAMA_EXCLUDE_LIST):
                continue

            qty_val = pd.to_numeric(qty, errors="coerce")
            if pd.notna(qty_val) and qty_val > 0:
                qty_round = round(float(qty_val), 0)
                rows.append(
                    {
                        "Date": date_only,
                        "Material": str(sku).strip() if sku is not None else None,
                        "Description": str(product).strip(),
                        "Qty": qty_round,
                    }
                )

    if not rows:
        return pd.DataFrame()

    return pd.DataFrame(rows)


def process_sakatama_file(
    file_bytes: bytes, sheet_name: str, month_set: set, cal_map: dict
) -> pd.DataFrame:
    out = extract_sakatama_production_data(file_bytes, sheet_name)
    if out.empty:
        return pd.DataFrame()

    # Enrichment from master data (by SKU)
    master_ref = load_fg_master_data(engine).copy()
    needed_cols = [
        "sku_code",
        "description",
        "kg_cb",
        "size",
        "speed",
        "country",
        "brand",
        "sub_brand",
        "category",
        "big_category",
        "house",
        "pack_format",
        "output",
    ]
    master_ref = master_ref[[c for c in needed_cols if c in master_ref.columns]].copy()
    master_ref["sku_code"] = master_ref["sku_code"].astype(str).str.strip()
    master_ref = master_ref.drop_duplicates(subset=["sku_code"], keep="first")

    out["Material"] = out["Material"].astype(str).str.strip()

    out = (
        out.merge(
            master_ref,
            how="left",
            left_on="Material",
            right_on="sku_code",
        )
        .drop(columns=["sku_code"], errors="ignore")
        .copy()
    )

    if "description" in out.columns:
        out["Description"] = out["Description"].where(
            out["Description"].notna() & (out["Description"].astype(str).str.strip() != ""),
            out["description"],
        )
        out = out.drop(columns=["description"], errors="ignore")

    if "size" in out.columns:
        out = out.rename(columns={"size": "Pack Size"})
    else:
        out["Pack Size"] = None

    if "kg_cb" in out.columns:
        out = out.rename(columns={"kg_cb": "Kg_TU"})
    else:
        out["Kg_TU"] = None

    if "speed" in out.columns:
        out = out.rename(columns={"speed": "Speed"})
    else:
        out["Speed"] = None

    out["Qty"] = pd.to_numeric(out["Qty"], errors="coerce").round(0)
    out["Kg_TU"] = pd.to_numeric(out["Kg_TU"], errors="coerce")
    out["Speed"] = pd.to_numeric(out["Speed"], errors="coerce")

    # Calculations
    out["Qty Bulk in KG"] = (out["Qty"] * out["Kg_TU"]).round(0)
    out["BIN"] = (out["Qty Bulk in KG"] / 750).round(0)

    out["Prod Hour"] = out.apply(
        lambda r: (r["Qty"] / r["Speed"])
        if pd.notna(r["Speed"]) and r["Speed"] not in [0, 0.0]
        else None,
        axis=1,
    )
    out["Days"] = out["Prod Hour"].apply(lambda x: (x / 24) if pd.notna(x) else None)

    # Remove duplicates at day/material/kg_tu level
    key_cols = ["Date", "Material", "Kg_TU"]
    out = out.drop_duplicates(subset=key_cols, keep="first")

    # Sorting + scheduling like East (single line)
    out = out.sort_values(["Date", "Material"], ascending=True).copy()
    out["_orig_date"] = pd.to_datetime(out["Date"], errors="coerce")

    sorted_rows = []
    unique_dates = sorted(out["_orig_date"].unique())
    last_material_prev_day = None

    for d in unique_dates:
        day_data = out[out["_orig_date"] == d].copy()

        if last_material_prev_day is not None:
            priority_df = day_data[day_data["Material"] == last_material_prev_day]
            others_df = day_data[day_data["Material"] != last_material_prev_day].sort_values(
                "Material", ascending=True
            )
            day_sorted = pd.concat([priority_df, others_df])
        else:
            day_sorted = day_data.sort_values("Material", ascending=True)

        if not day_sorted.empty:
            last_material_prev_day = day_sorted.iloc[-1]["Material"]

        sorted_rows.append(day_sorted)

    out = pd.concat(sorted_rows).reset_index(drop=True)

    # Time calculations
    time_starts = []
    time_finishes = []

    for idx, row in out.iterrows():
        current_date_6am = pd.Timestamp(row["_orig_date"]) + pd.Timedelta(hours=6)

        if idx == 0:
            time_start = pd.Timestamp(row["_orig_date"]) + pd.Timedelta(hours=7)
        else:
            prev_finish = time_finishes[-1]
            time_start = prev_finish if prev_finish > current_date_6am else current_date_6am

        days_value = row["Days"] if pd.notna(row["Days"]) else 0
        time_finish = time_start + pd.Timedelta(days=days_value)

        time_starts.append(time_start)
        time_finishes.append(time_finish)

    out["Time Start"] = time_starts
    out["Time Finish"] = time_finishes

    # Filter by month set (M0..M2)
    out = out[out["Time Finish"].dt.month.isin(month_set)].copy()
    if out.empty:
        return pd.DataFrame()

    out = out.sort_values("Time Start", ascending=True).reset_index(drop=True)

    out["Release Time"] = out["Time Finish"].apply(calc_release_time)
    out["Release Time"] = pd.to_datetime(out["Release Time"], errors="coerce").dt.date
    out["Release wk"] = out["Release Time"].map(cal_map)

    # Date to Mon-YY string for final output
    out["Date"] = pd.to_datetime(out["Date"], errors="coerce").dt.strftime("%b-%y")

    out = out.rename(
        columns={
            "Material": "SAP Article",
            "Qty": "Qty (Ctn)",
            "Qty Bulk in KG": "Qty Bulk (kg)",
            "Release wk": "Release Week",
        }
    )

    if "Release Time" in out.columns and "Release Week" in out.columns:

        def rel_ident_fmt(x):
            if pd.notna(x):
                return f"{x.day}{x.month}{x.year}"
            return ""

        rel_ident = out["Release Time"].apply(rel_ident_fmt)
        idx = out.columns.get_loc("Release Week")
        out.insert(idx + 1, "Release Ident", rel_ident)

    # Ensure Line column exists
    out.insert(0, "Line", SAKATAMA_LINE)

    # Final selection of columns
    final_cols_with_time = [
        "Line",
        "Date",
        "SAP Article",
        "Description",
        "Pack Size",
        "Kg_TU",
        "Qty (Ctn)",
        "Qty Bulk (kg)",
        "BIN",
        "Time Start",
        "Time Finish",
        "Release Time",
        "Release Week",
        "Release Ident",
        "country",
        "brand",
        "sub_brand",
        "category",
        "big_category",
        "house",
        "pack_format",
    ]
    out = out[[c for c in final_cols_with_time if c in out.columns]].copy()

    return out


def create_sakatama_excel_download(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=f"Line_{SAKATAMA_LINE}", index=False)
        df.to_excel(writer, sheet_name="All_Sakatama", index=False)
    output.seek(0)
    return output.getvalue()


# -------------------------
# UI render functions (Tabs)
# -------------------------
def render_west():
    # Month inputs (unique keys because East has identical widgets)
    c1m, c2m, c3m = st.columns(3)
    with c1m:
        m0 = st.number_input(
            "M0 Month (1-12)",
            min_value=1,
            max_value=12,
            value=2,
            step=1,
            key="west_m0",
        )
    with c2m:
        m1 = ((m0 - 1 + 1) % 12) + 1
        st.text_input("M1", value=str(m1), disabled=True, key="west_m1")
    with c3m:
        m2 = ((m0 - 1 + 2) % 12) + 1
        st.text_input("M2", value=str(m2), disabled=True, key="west_m2")
    month_set = {int(m0), int(m1), int(m2)}
    st.markdown("---")

    uploaded = st.file_uploader("Upload Excel (.xlsx)", type=["xlsx"], key="west_upload")
    if not uploaded:
        st.caption("Upload your file to start the process.")
        return

    xls = pd.ExcelFile(uploaded, engine="openpyxl")
    sheet_options = xls.sheet_names
    selected_sheets = st.multiselect(
        "Pilih sheet yang ingin diproses:",
        options=sheet_options,
        default=[],
        help="Pilih satu atau lebih sheet untuk diproses",
        key="west_selected_sheets",
    )

    if st.button("Process Selected Sheets", disabled=not selected_sheets, key="west_process_btn"):
        with st.spinner("Processing sheets..."):
            results = {}
            report = []
            for sh in selected_sheets:
                try:
                    df_out, status = process_sheet(uploaded, sh, month_set)
                    rows = 0 if df_out is None else len(df_out)
                    report.append((sh, status, rows))
                    if df_out is not None and not df_out.empty:
                        df_out.insert(0, "Region", "West")
                        if "Line" in df_out.columns:    
                            df_out = df_out.rename(columns={"Line": "Date"})
                        region_idx = df_out.columns.get_loc("Region")
                        df_out.insert(region_idx + 1, "Line", sh)
                        df_out = df_out.rename(
                            columns={
                                "Release wk": "Release Week",
                                "Time_Finish": "Time Finish",
                                "Release time": "Release Time",
                            }
                        )
                        if "Release Time" in df_out.columns and "Release Week" in df_out.columns:

                            def rel_ident_fmt(x):
                                if pd.notna(x):
                                    return f"{x.day}{x.month}{x.year}"
                                return ""

                            rel_ident = df_out["Release Time"].apply(rel_ident_fmt)
                            idx = df_out.columns.get_loc("Release Week")
                            df_out.insert(idx + 1, "Release Ident", rel_ident)

                        # Samakan nama kolom agar konsisten di All_West
                        df_out = df_out.rename(columns={"Qty Bulk in KG": "Qty Bulk (kg)"})

                        # Normalize Qty header variants (e.g., 'Qty (ctn)') -> 'Qty (Ctn)'
                        def _norm_key_local(s: str) -> str:
                            return re.sub(r"[^a-z0-9]", "", str(s or "").lower())

                        col_map_local = {_norm_key_local(c): c for c in df_out.columns}
                        if "qtyctn" in col_map_local:
                            orig = col_map_local["qtyctn"]
                            if orig != "Qty (Ctn)":
                                df_out = df_out.rename(columns={orig: "Qty (Ctn)"})

                        # Remove enrichment columns for West outputs (keep them only in Combined)
                        for _c in [
                            "country",
                            "brand",
                            "sub_brand",
                            "category",
                            "big_category",
                            "house",
                            "pack_format",
                            "machine_1",
                        ]:
                            if _c in df_out.columns:
                                df_out = df_out.drop(columns=[_c])

                        results[sh] = df_out
                except Exception as e:
                    report.append((sh, f"ERROR: {e}", 0))

            rep_df = pd.DataFrame(report, columns=["Sheet", "Status", "Rows"])

            if not results:
                st.error("No sheets were successfully processed (all skipped or error).")
                st.dataframe(rep_df, use_container_width=True)
                return

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                for sh, df_out in results.items():
                    df_out.to_excel(writer, sheet_name=sh, index=False)
                if results:
                    all_west_df = pd.concat(results.values(), ignore_index=True)
                    all_west_df.to_excel(writer, sheet_name="All_West", index=False)
            output.seek(0)

        st.success(f"Done! Sheets processed: {len(results)} / {len(selected_sheets)}")

        st.subheader("Preview Data per Sheet")
        for sh, df_out in results.items():
            with st.expander(f"📄 {sh} ({len(df_out)} rows)", expanded=False):
                st.dataframe(df_out, use_container_width=True)

        # Format nama file pakai singkatan bulan M0-M2
        month_names = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]
        m0_name = month_names[(int(m0) - 1) % 12]
        file_name = f"DPS WEST {m0_name} Output.xlsx"

        st.download_button(
            "Download Output (Excel)",
            data=output,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="west_download_btn",
        )


def render_east():
    c1m, c2m, c3m = st.columns(3)
    with c1m:
        m0 = st.number_input(
            "M0 Month (1-12)",
            min_value=1,
            max_value=12,
            value=2,
            step=1,
            key="east_m0",
        )
    with c2m:
        m1 = ((m0 - 1 + 1) % 12) + 1
        st.text_input("M1", value=str(m1), disabled=True, key="east_m1")
    with c3m:
        m2 = ((m0 - 1 + 2) % 12) + 1
        st.text_input("M2", value=str(m2), disabled=True, key="east_m2")
    month_set = {int(m0), int(m1), int(m2)}
    st.markdown("---")

    uploaded = st.file_uploader("Upload Excel (.xlsx)", type=["xlsx"], key="east_upload")
    if not uploaded:
        st.caption("Upload your file to start the process.")
        return

    try:
        @st.cache_data
        def get_sheet_names(file_bytes):
            """Cache sheet names to avoid re-reading Excel file"""
            return pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl").sheet_names

        file_bytes = uploaded.getvalue()
        sheet_names = get_sheet_names(file_bytes)
        st.markdown("---")

        # Multiselect untuk pilih beberapa sheet
        selected_sheets = st.multiselect(
            "Pilih sheet yang ingin diproses:",
            options=sheet_names,
            default=[],
            help="Pilih satu atau lebih sheet untuk diproses",
            key="east_selected_sheets",
        )

        if st.button("Process Selected Sheets", disabled=not selected_sheets, key="east_process_btn"):
            # Gabungkan semua hasil line dari semua sheet berdasarkan nama line
            combined_line_dfs = {}
            error_report = []
            for selected_sheet in selected_sheets:
                with st.spinner(f"Reading sheet '{selected_sheet}'..."):
                    raw = pd.read_excel(
                        io.BytesIO(file_bytes),
                        sheet_name=selected_sheet,
                        header=None,
                        engine="openpyxl",
                    )
                    marker = "Total SH Production"
                    cut_row = None
                    for idx, row in raw.iterrows():
                        if row.astype(str).str.contains(marker, case=False, na=False).any():
                            cut_row = idx
                            break
                    if cut_row is not None:
                        raw = raw.iloc[:cut_row, :].copy()

                with st.spinner(f"Validating & processing '{selected_sheet}'..."):
                    is_valid, error_message = validate_east_sheet_format(raw)
                    if not is_valid:
                        error_report.append((selected_sheet, error_message))
                        continue
                    try:
                        line_dfs = process_east_file(raw, engine, month_set, CAL_MAP)
                        if not line_dfs:
                            error_report.append((selected_sheet, "No data found after processing."))
                        else:
                            for line, df in line_dfs.items():
                                # Keep df columns as produced by process_east_file (target columns)
                                if line not in combined_line_dfs:
                                    combined_line_dfs[line] = [df]
                                else:
                                    combined_line_dfs[line].append(df)
                    except Exception as e:
                        error_report.append((selected_sheet, f"Error processing file: {str(e)}"))

            # Gabungkan DataFrame per line
            final_line_dfs = {}
            for line, df_list in combined_line_dfs.items():
                final_line_dfs[line] = pd.concat(df_list, ignore_index=True)

            if not final_line_dfs:
                st.error("No data found in any selected sheet.")
                if error_report:
                    st.dataframe(
                        pd.DataFrame(error_report, columns=["Sheet", "Error"]),
                        use_container_width=True,
                    )
                return

            st.success(
                f"Done! Sheets processed: {len(final_line_dfs)} line(s) from {len(selected_sheets)} sheet(s)"
            )

            st.markdown("---")
            st.subheader("Download Output")
            # Remove enrichment columns for East outputs (they are used only in Combined)
            for lk, ldf in list(final_line_dfs.items()):
                for _c in [
                    "country",
                    "brand",
                    "sub_brand",
                    "category",
                    "big_category",
                    "house",
                    "pack_format",
                    "machine_1",
                ]:
                    if _c in ldf.columns:
                        ldf = ldf.drop(columns=[_c])
                final_line_dfs[lk] = ldf

            excel_data = create_east_excel_download(final_line_dfs)
            st.download_button(
                label="Download Excel File",
                data=excel_data,
                file_name="DPS East Output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="east_download_btn",
            )

            st.markdown("---")
            st.subheader("Preview Sheets")
            line_tabs = st.tabs([f"Line_{k}" for k in sorted(final_line_dfs.keys())])
            for tab, k in zip(line_tabs, sorted(final_line_dfs.keys())):
                with tab:
                    df = final_line_dfs[k]
                    st.write(f"**Total rows:** {len(df)}")
                    st.dataframe(df, use_container_width=True, height=400)

            if error_report:
                st.markdown("---")
                st.subheader("Error Report")
                st.dataframe(
                    pd.DataFrame(error_report, columns=["Sheet", "Error"]),
                    use_container_width=True,
                )

    except Exception as e:
        st.error(f"Error reading file: {str(e)}")
        st.exception(e)


def render_sakatama():
    c1m, c2m, c3m = st.columns(3)
    with c1m:
        m0 = st.number_input(
            "M0 Month (1-12)",
            min_value=1,
            max_value=12,
            value=2,
            step=1,
            key="sakatama_m0",
        )
    with c2m:
        m1 = ((m0 - 1 + 1) % 12) + 1
        st.text_input("M1", value=str(m1), disabled=True, key="sakatama_m1")
    with c3m:
        m2 = ((m0 - 1 + 2) % 12) + 1
        st.text_input("M2", value=str(m2), disabled=True, key="sakatama_m2")
    month_set = {int(m0), int(m1), int(m2)}
    st.markdown("---")

    uploaded = st.file_uploader("Upload Excel (.xlsx)", type=["xlsx"], key="sakatama_upload")
    if not uploaded:
        st.caption("Upload your file to start the process.")
        return

    try:
        @st.cache_data
        def get_sheet_names(file_bytes):
            """Cache sheet names to avoid re-reading Excel file"""
            return pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl").sheet_names

        file_bytes = uploaded.getvalue()
        sheet_names = get_sheet_names(file_bytes)
        st.markdown("---")

        selected_sheets = st.multiselect(
            "Pilih sheet yang ingin diproses:",
            options=sheet_names,
            default=[],
            help="Pilih satu atau lebih sheet untuk diproses",
            key="sakatama_selected_sheets",
        )

        if st.button(
            "Process Selected Sheets", disabled=not selected_sheets, key="sakatama_process_btn"
        ):
            all_dfs = []
            error_report = []

            for selected_sheet in selected_sheets:
                with st.spinner(f"Reading & processing '{selected_sheet}'..."):
                    try:
                        df = process_sakatama_file(file_bytes, selected_sheet, month_set, CAL_MAP)
                        if df is None or df.empty:
                            error_report.append((selected_sheet, "No data found after processing."))
                        else:
                            all_dfs.append(df)
                    except Exception as e:
                        error_report.append((selected_sheet, f"Error processing file: {str(e)}"))

            if not all_dfs:
                st.error("No data found in any selected sheet.")
                if error_report:
                    st.dataframe(
                        pd.DataFrame(error_report, columns=["Sheet", "Error"]),
                        use_container_width=True,
                    )
                return

            final_df = pd.concat(all_dfs, ignore_index=True)

            st.success(f"Done! Sheets processed: {len(all_dfs)} sheet(s)")

            st.markdown("---")
            st.subheader("Download Output")

            # Remove enrichment columns for Sakatama output
            for _c in [
                "country",
                "brand",
                "sub_brand",
                "category",
                "big_category",
                "house",
                "pack_format",
                "machine_1",
            ]:
                if _c in final_df.columns:
                    final_df = final_df.drop(columns=[_c])

            excel_data = create_sakatama_excel_download(final_df)
            st.download_button(
                label="Download Excel File",
                data=excel_data,
                file_name="DPS Sakatama Output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="sakatama_download_btn",
            )

            st.markdown("---")
            st.subheader("Preview Sheets")
            line_tabs = st.tabs([f"Line_{SAKATAMA_LINE}"])
            with line_tabs[0]:
                st.write(f"**Total rows:** {len(final_df)}")
                st.dataframe(final_df, use_container_width=True, height=400)

            if error_report:
                st.markdown("---")
                st.subheader("Error Report")
                st.dataframe(
                    pd.DataFrame(error_report, columns=["Sheet", "Error"]),
                    use_container_width=True,
                )

    except Exception as e:
        st.error(f"Error reading file: {str(e)}")
        st.exception(e)


def render_combined():
    col1, col2, col3 = st.columns(3)
    with col1:
        file_west = st.file_uploader(
            "Upload West Excel (.xlsx)",
            type=["xlsx"],
            key="combined_west",
        )
    with col2:
        file_east = st.file_uploader(
            "Upload East Excel (.xlsx)",
            type=["xlsx"],
            key="combined_east",
        )
        
    with col3:
        file_sakatama = st.file_uploader(
            "Upload Sakatama Excel (.xlsx)",
            type=["xlsx"],
            key="combined_sakatama",
        )

    # Tombol hanya aktif jika kedua file sudah ada
    can_process = file_west is not None and file_east is not None and file_sakatama is not None
    start = st.button("Start Processing", disabled=not can_process, key="combined_start")

    if not start:
        return

    try:
        # Validasi sheet All_West & All_East
        xls_west = pd.ExcelFile(file_west, engine="openpyxl")
        xls_east = pd.ExcelFile(file_east, engine="openpyxl")
        xls_sakatama = pd.ExcelFile(file_sakatama, engine="openpyxl")
        sheetnames_west = [s.strip().lower() for s in xls_west.sheet_names]
        sheetnames_east = [s.strip().lower() for s in xls_east.sheet_names]
        sheetnames_sakatama = [s.strip().lower() for s in xls_sakatama.sheet_names]

        if "all_west" not in sheetnames_west or "all_east" not in sheetnames_east:
            st.error("Pastikan file West punya sheet 'All_West' dan file East punya sheet 'All_East'.")
            return
        
        if "all_sakatama" not in sheetnames_sakatama:
            st.error("Pastikan file Sakatama punya sheet 'All_Sakatama'.")
            return

        # Baca sheet
        df_west = pd.read_excel(file_west, sheet_name="All_West", header=0, engine="openpyxl")
        df_east = pd.read_excel(file_east, sheet_name="All_East", header=0, engine="openpyxl")
        df_sakatama = pd.read_excel(file_sakatama, sheet_name="All_Sakatama", header=0, engine="openpyxl")

        # Target combined column order
        TARGET_COMBINED_COLS = [
            "Region",
            "Line",
            "SAP Article",
            "Description",
            "Pack Size",
            "Kg_TU",
            "Qty (Ctn)",
            "Qty Bulk (kg)",
            "BIN",
            "Time Start",
            "Time Finish",
            "Release Time",
            "Release Week",
            "Release Ident",
            "Country",
            "Brand",
            "Sub Brand",
            "Category",
            "Big Category",
            "House",
            "Pack Format",
            "Ouput",
        ]

        def process_combined_file(df: pd.DataFrame, region_label: str) -> pd.DataFrame:
            """Normalize columns from an All_West/All_East file and enrich from fg_master_data (MASTER_MAP)."""

            def norm_key(s: str) -> str:
                s = str(s or "").lower()
                return re.sub(r"[^a-z0-9]", "", s)

            col_map = {norm_key(c): c for c in df.columns}

            out = pd.DataFrame()

            # Core columns to take from the sheet
            sheet_fields = [
                "Line",
                "SAP Article",
                "Description",
                "Pack Size",
                "Kg_TU",
                "Qty (Ctn)",
                "Qty Bulk (kg)",
                "BIN",
                "Time Start",
                "Time Finish",
                "Release Time",
                "Release Week",
            ]

            for f in sheet_fields:
                k = norm_key(f)
                out[f] = df[col_map[k]] if k in col_map else None

            # Ensure datetime for Time Start/Finish/Release Time
            for t in ["Time Start", "Time Finish", "Release Time"]:
                if t in out.columns:
                    out[t] = pd.to_datetime(out[t], errors="coerce")

            # Compute Release Ident
            if "Release Time" in out.columns:

                def rel_ident_fmt(x):
                    if pd.notna(x):
                        return f"{x.day}{x.month}{x.year}"
                    return None

                out["Release Ident"] = out["Release Time"].apply(rel_ident_fmt)
            else:
                out["Release Ident"] = None

            # Enrich from MASTER_MAP (fg_master_data) using SAP Article/material code
            def enrich_row(mat):
                k = str(mat).strip()
                info = MASTER_MAP.get(k, {}) if k else {}
                return {
                    "Country": info.get("country"),
                    "Brand": info.get("brand"),
                    "Sub Brand": info.get("sub_brand"),
                    "Category": info.get("category"),
                    "Big Category": info.get("big_category"),
                    "House": info.get("house"),
                    "Pack Format": info.get("pack_format"),
                    "Ouput": info.get("output"),
                }

            enrich_df = pd.DataFrame(list(map(enrich_row, out["SAP Article"].fillna(""))))
            out = pd.concat([out.reset_index(drop=True), enrich_df.reset_index(drop=True)], axis=1)

            # (Opsional) isi Description kosong dari master
            if "Description" in out.columns:

                def fill_desc(row):
                    d = row.get("Description")
                    if pd.isna(d) or str(d).strip() == "":
                        k = str(row.get("SAP Article") or "").strip()
                        return MASTER_MAP.get(k, {}).get("description")
                    return d

                out["Description"] = out.apply(fill_desc, axis=1)

            # Region column set from the file source
            out.insert(0, "Region", region_label)

            # Normalize Time Start/Finish/Release Time to date (no time)
            for t in ["Time Start", "Time Finish", "Release Time"]:
                if t in out.columns:
                    out[t] = pd.to_datetime(out[t], errors="coerce").dt.date

            # Ensure final column order and missing columns
            final = pd.DataFrame()
            existing_map = {c.lower().strip(): c for c in out.columns}
            for tc in TARGET_COMBINED_COLS:
                key = tc.lower().strip()
                final[tc] = out[existing_map[key]] if key in existing_map else None

            return final

        df_west_sel = process_combined_file(df_west, "West")
        df_east_sel = process_combined_file(df_east, "East")
        df_sakatama_sel = process_combined_file(df_sakatama, "Sakatama")
        df_combined = pd.concat([df_west_sel, df_east_sel, df_sakatama_sel], ignore_index=True)

        st.success(
            f"Data digabungkan: {len(df_west_sel)} baris dari West, "
            f"{len(df_east_sel)} baris dari East, "
            f"{len(df_sakatama_sel)} baris dari Sakatama, "
            f"total {len(df_combined)} baris."
        )
        st.markdown("---")
        st.subheader("Preview Combined Data")
        st.dataframe(df_combined, use_container_width=True, height=400)

        # Download button
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_combined.to_excel(writer, sheet_name="Combined_DPS", index=False)
        output.seek(0)
        st.download_button(
            label="Download Combined Excel",
            data=output,
            file_name="DPS Total East West Output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="combined_download_btn",
        )

    except Exception as e:
        st.error(f"Gagal membaca file: {str(e)}")
        return


# -------------------------
# Tabs (replaces radio)
# -------------------------
tab_west, tab_east, tab_sakatama, tab_combined = st.tabs(["West", "East", "Sakatama", "Combined"])
with tab_west:
    render_west()
with tab_east:
    render_east()
with tab_sakatama:
    render_sakatama()
with tab_combined:
    render_combined()