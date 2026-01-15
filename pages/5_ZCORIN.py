import io
import os
import pandas as pd
import streamlit as st
from sqlalchemy import create_engine, text
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="ZCORIN Cleaner", layout="wide")
st.title("ZCORIN Cleaner")

# =========================
# DB ENGINE (Neon)
# =========================
@st.cache_resource
def get_engine():
    p = st.secrets["postgres"]
    url = (
        f"postgresql+psycopg2://{p['user']}:{p['password']}"
        f"@{p['host']}:{p['port']}/{p['database']}"
        f"?sslmode=require"
    )
    return create_engine(url, pool_pre_ping=True)

engine = get_engine()

def load_conversion_map():
    """material -> pcs_cb (numeric)"""
    sql = text("SELECT material, pcs_cb FROM zcorin_converter")
    with engine.connect() as conn:
        df = pd.read_sql(sql, conn)

    df["material"] = df["material"].astype(str).str.strip()
    df["pcs_cb"] = pd.to_numeric(df["pcs_cb"], errors="coerce")
    return dict(zip(df["material"], df["pcs_cb"]))

def parse_date_series(s: pd.Series) -> pd.Series:
    """
    Convert date-like strings from SAP/Excel to real datetime.
    Primary guess: mm/dd/yyyy (common in exports),
    fallback: generic parse.
    """
    # try strict US-style first
    a = pd.to_datetime(s, format="%m/%d/%Y", errors="coerce")
    # fallback parse
    b = pd.to_datetime(s, errors="coerce", dayfirst=False)
    return a.combine_first(b)

# =========================
# UI
# =========================
uploaded = st.file_uploader("Upload file ZCORIN (.xlsx)", type=["xlsx"])
start_time = st.date_input("Start Time (manual input user)", value=None)

if not uploaded:
    st.info("Upload file first to start.")
    st.stop()

if not start_time:
    st.warning("Fill in Start Time first (date input).")
    st.stop()

# =========================
# PROCESS
# =========================
if st.button("Start process ZCORIN"):
    with st.spinner("Processing..."):
        # Read
        df = pd.read_excel(uploaded, sheet_name="Sheet1", engine="openpyxl")

        # Column positions: B & M
        storage_col = df.columns[1]   # col B
        unit_col = df.columns[12]     # col M

        # Filter
        df_f = df[
            (df[storage_col].isin([1, 6]) | df[storage_col].isna()) &
            (df[unit_col].astype(str).str.strip().str.upper() == "PC")
        ].copy()

        # Sort Storage Location: blank -> 1 -> 6
        def storage_sort_key(x):
            if pd.isna(x) or str(x).strip() == "":
                return 0
            if str(x).strip() == "1":
                return 1
            if str(x).strip() == "6":
                return 2
            return 99

        df_f["_storage_sort"] = df_f[storage_col].apply(storage_sort_key)
        df_f = df_f.sort_values("_storage_sort").drop(columns="_storage_sort")

        # Required columns
        required_cols = [
            "Material",
            "Unrestricted",
            "Blocked",
            "Qual. Inspection",
            "Transfer",
            "Returns(Blocked)",
            "In Transit-Receivi",
            "SLED/BBD",
            "Manuf. Dte",
        ]
        missing = [c for c in required_cols if c not in df_f.columns]
        if missing:
            st.error(f"Kolom ini tidak ditemukan di file: {missing}")
            st.stop()

        # ---- FIX: convert date columns to real datetime (so formulas work in Excel) ----
        df_f["SLED/BBD"] = parse_date_series(df_f["SLED/BBD"])
        df_f["Manuf. Dte"] = parse_date_series(df_f["Manuf. Dte"])

        # Add Start Time as DATE (no time)
        df_f["Start Time"] = start_time

        # Conversion from DB (pcs_cb)
        conv_map = load_conversion_map()
        df_f["Conversion"] = df_f["Material"].astype(str).str.strip().map(conv_map)

        # Add requested columns (placeholders; formulas injected later)
        df_f["Unrestricted_vis"] = None
        df_f["Blocked_vis"] = None
        df_f["Qual. Inspection_vis"] = None
        df_f["Transfer_vis"] = None
        df_f["Returns(Blocked)_vis"] = None
        df_f["Unit_vis"] = "Ctn"
        df_f["MRP Controller_vis"] = ""
        df_f["Vendor Batch_vis"] = ""
        df_f["In Transit-Receivi_vis"] = None
        df_f["Total_vis"] = None
        df_f["Shelf Life"] = None
        df_f["Total Shelf life (years)"] = None
        df_f["Remaining Shelf life (%)"] = None
        df_f["Aging (month)"] = None

        # Save to Excel in memory
        base_name = os.path.splitext(uploaded.name)[0]
        out_name = f"{base_name}_vis.xlsx"

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df_f.to_excel(writer, index=False, sheet_name="Output")
        buffer.seek(0)

        # Inject formulas using openpyxl
        wb = load_workbook(buffer)
        ws = wb["Output"]

        # Map header -> column index (1-based)
        headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

        def cell(row, col_name):
            return f"{get_column_letter(headers[col_name])}{row}"

        # Format date columns to show date only
        for col_name in ["SLED/BBD", "Manuf. Dte", "Start Time"]:
            cidx = headers.get(col_name)
            if cidx:
                for rr in range(2, ws.max_row + 1):
                    ws.cell(row=rr, column=cidx).number_format = "dd-mm-yyyy"

        last_row = ws.max_row
        for r in range(2, last_row + 1):
            conv = cell(r, "Conversion")

            ws[cell(r, "Unrestricted_vis")] = f'=IFERROR({cell(r,"Unrestricted")}/{conv},"")'
            ws[cell(r, "Blocked_vis")] = f'=IFERROR({cell(r,"Blocked")}/{conv},"")'
            ws[cell(r, "Qual. Inspection_vis")] = f'=IFERROR({cell(r,"Qual. Inspection")}/{conv},"")'
            ws[cell(r, "Transfer_vis")] = f'=IFERROR({cell(r,"Transfer")}/{conv},"")'
            ws[cell(r, "Returns(Blocked)_vis")] = f'=IFERROR({cell(r,"Returns(Blocked)")}/{conv},"")'
            ws[cell(r, "In Transit-Receivi_vis")] = f'=IFERROR({cell(r,"In Transit-Receivi")}/{conv},"")'

            # Total = Unrestricted_vis + Qual. Inspection_vis + In Transit-Receivi_vis
            ws[cell(r, "Total_vis")] = (
                f'=IFERROR({cell(r,"Unrestricted_vis")}+{cell(r,"Qual. Inspection_vis")}+{cell(r,"In Transit-Receivi_vis")},"")'
            )

            # Date-based formulas (rounded 2 decimals)
            ws[cell(r, "Shelf Life")] = (
                f'=IFERROR(ROUND(({cell(r,"SLED/BBD")}-{cell(r,"Start Time")})/360, 2), "")'
            )

            ws[cell(r, "Total Shelf life (years)")] = (
                f'=IFERROR(ROUND(({cell(r,"SLED/BBD")}-{cell(r,"Manuf. Dte")})/360, 2), "")'
            )

            # Remaining (%) = (Shelf Life / Total Shelf life) * 100
            ws[cell(r, "Remaining Shelf life (%)")] = (
                f'=IFERROR(ROUND(({cell(r,"Shelf Life")}/{cell(r,"Total Shelf life (years)")})*100, 2), "")'
            )

            # show percent sign after number (numeric stays numeric)
            pct_cell = ws[cell(r, "Remaining Shelf life (%)")]
            pct_cell.number_format = '0.00"%"'

            ws[cell(r, "Aging (month)")] = (
                f'=IFERROR(ROUND(({cell(r,"Start Time")}-{cell(r,"Manuf. Dte")})/30, 2), "")'
            )

        out_bytes = io.BytesIO()
        wb.save(out_bytes)
        out_bytes.seek(0)

    st.success("Cleansing Done!")
    st.download_button(
        "Download Output (Excel)",
        data=out_bytes,
        file_name=out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )