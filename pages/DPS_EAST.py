import os
import re
import pandas as pd
from sqlalchemy import create_engine, text
from openpyxl import load_workbook

# =========================
# PATHS
# =========================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)

FILE_PATH = os.path.join(
    PROJECT_ROOT,
    "src", "east",
    "PP_SH2_Jan - Apr 26_upd 19 Jan deliv 19jan_rofo 20 Dec (version 1).xlsx"
)

SECRETS_PATH = os.path.join(PROJECT_ROOT, ".streamlit", "secrets.toml")

SHEET_NAME = "Daily_Jan-Feb"
OUTPUT_XLSX = os.path.join(PROJECT_ROOT, "DPS_East_Daily_Jan-Feb_VERTICAL_ENRICHED.xlsx")

# =========================
# EXCEL COORDINATES (0-based)
# =========================
DATE_ROW_IDX = 8
DATE_START_COL = 24   # Y
DATE_END_COL = 93     # CP

COL_MATERIAL = 5      # F
COL_DESC = 6          # G
COL_KG_TU = 9         # J (kg/cb)
COL_LINE = 10         # K

VALID_LINES = {"AB", "CD", "GH", "JK", "TU", "VW", "XY"}

# =========================
# READ SECRETS.TOML (without streamlit)
# =========================
def load_secrets_toml(path: str) -> dict:
    if not os.path.exists(path):
        raise FileNotFoundError(f"secrets.toml not found at: {path}")

    try:
        import tomllib  # py>=3.11
        with open(path, "rb") as f:
            return tomllib.load(f)
    except ModuleNotFoundError:
        import toml  # pip install toml
        return toml.load(path)

def get_engine_from_secrets(secrets: dict):
    if "postgres" not in secrets:
        raise KeyError("Section [postgres] not found in secrets.toml")

    p = secrets["postgres"]
    required = ["host", "port", "database", "user", "password"]
    missing = [k for k in required if k not in p]
    if missing:
        raise KeyError(f"Missing keys in [postgres]: {missing}")

    url = (
        f"postgresql+psycopg2://{p['user']}:{p['password']}"
        f"@{p['host']}:{p['port']}/{p['database']}"
        f"?sslmode=require"
    )
    return create_engine(url, pool_pre_ping=True)

# =========================
# DB: load enrichment
# =========================
def load_zcorin_converter(engine) -> pd.DataFrame:
    sql = text("""
        SELECT material, country, brand, big_category, house, pack_format, machine_1
        FROM zcorin_converter
    """)
    with engine.connect() as conn:
        df = pd.read_sql(sql, conn)
    df["material"] = df["material"].astype(str).str.strip()
    return df

# =========================
# Find cut row using openpyxl (robust)
# =========================
def find_cut_row_by_marker_xlsx(xlsx_path: str, sheet_name: str, marker_text: str) -> int | None:
    """
    Returns 1-based row number of the first row containing marker_text (case-insensitive),
    searching ALL cells. If not found, returns None.
    """
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")

    ws = wb[sheet_name]
    pat = re.compile(re.escape(marker_text), re.IGNORECASE)

    for row in ws.iter_rows(values_only=True):
        # row is a tuple of values
        for v in row:
            if v is None:
                continue
            s = str(v)
            if pat.search(s):
                # openpyxl gives row index via ws.iter_rows doesn't directly expose index,
                # so we compute using ws._current_row is not reliable in read_only.
                # Instead, iterate with enumerate over iter_rows.
                pass

    # Do a reliable enumerate iteration to get the row index:
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for v in row:
            if v is None:
                continue
            if pat.search(str(v)):
                wb.close()
                return r_idx

    wb.close()
    return None

# =========================
# MAIN
# =========================
def main():
    # ---- DB connect via secrets.toml ----
    secrets = load_secrets_toml(SECRETS_PATH)
    engine = get_engine_from_secrets(secrets)
    conv = load_zcorin_converter(engine)

    # ---- CUT logic (by marker) ----
    marker = "Total SH Production"
    cut_row = find_cut_row_by_marker_xlsx(FILE_PATH, SHEET_NAME, marker)

    if cut_row is not None:
        # Keep rows strictly ABOVE marker row
        nrows = cut_row - 1
        print(f"[CUT] Found '{marker}' at Excel row {cut_row}. Reading only first {nrows} rows.")
    else:
        nrows = None
        print(f"[CUT] Marker '{marker}' not found. Reading full sheet.")

    # ---- Read raw excel (already cut) ----
    raw = pd.read_excel(
        FILE_PATH,
        sheet_name=SHEET_NAME,
        header=None,
        nrows=nrows,
        engine="openpyxl"
    )

    # ---- Dates from Y9:CP9 ----
    date_cells = raw.iloc[DATE_ROW_IDX, DATE_START_COL:DATE_END_COL + 1]
    dates = pd.to_datetime(date_cells, errors="coerce")

    valid_date_mask = dates.notna()
    date_cols_idx = [DATE_START_COL + i for i, ok in enumerate(valid_date_mask.tolist()) if ok]
    date_vals = dates[valid_date_mask].dt.date.tolist()

    if not date_cols_idx:
        raise ValueError("No valid date headers found in Y9:CP9.")

    # ---- Filter item rows by Line (K) ----
    df_items = raw.copy()
    df_items[COL_LINE] = df_items[COL_LINE].astype(str).str.strip().str.upper()
    df_items = df_items[df_items[COL_LINE].isin(VALID_LINES)].copy()

    if df_items.empty:
        raise ValueError("No rows found with Line in AB/CD/GH/JK/TU/VW/XY.")

    # ---- Wide table ----
    keep_cols = [COL_MATERIAL, COL_DESC, COL_KG_TU, COL_LINE] + date_cols_idx
    df_wide = df_items.iloc[:, keep_cols].copy()
    df_wide.columns = ["Material", "Description", "Kg_TU", "Line"] + [str(d) for d in date_vals]

    df_wide["Material"] = df_wide["Material"].astype(str).str.strip()
    df_wide["Description"] = df_wide["Description"].astype(str).str.strip()
    df_wide["Line"] = df_wide["Line"].astype(str).str.strip().str.upper()
    df_wide["Kg_TU"] = pd.to_numeric(df_wide["Kg_TU"], errors="coerce")

    # remove blank materials
    df_wide = df_wide[(df_wide["Material"] != "") & (df_wide["Material"].str.lower() != "nan")].copy()

    # ---- Unpivot to vertical ----
    out = df_wide.melt(
        id_vars=["Material", "Description", "Kg_TU", "Line"],
        var_name="Date",
        value_name="Qty"
    )

    out["Date"] = pd.to_datetime(out["Date"], errors="coerce").dt.date
    out["Qty"] = pd.to_numeric(out["Qty"], errors="coerce")

    # ✅ keep only Qty > 0
    out = out[out["Qty"].fillna(0) > 0].copy()

    # ---- Enrich from DB ----
    out["Material"] = out["Material"].astype(str).str.strip()
    conv["material"] = conv["material"].astype(str).str.strip()

    out = out.merge(conv, how="left", left_on="Material", right_on="material").drop(columns=["material"], errors="ignore")

    # ---- Final order + sort ----
    final_cols = [
        "Date", "Material", "Description", "Kg_TU", "Line", "Qty",
        "country", "brand", "big_category", "house", "pack_format", "machine_1"
    ]
    out = out[[c for c in final_cols if c in out.columns]].copy()
    out = out.sort_values(["Date", "Line", "Material"], ascending=True)

    # ---- Save ----
    out.to_excel(OUTPUT_XLSX, index=False)

    print("Done!")
    print("Rows:", len(out))
    print("Saved:", OUTPUT_XLSX)

if __name__ == "__main__":
    main()